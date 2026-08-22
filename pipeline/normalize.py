"""Normalize DPI statewide CSVs into per-district JSON.

Two layers of builders share the same per-topic metric extraction:

* ``BUILDERS[topic](frames)`` — all-students metrics for EVERY Wisconsin
  district (plus statewide as code 0000):
  ``{dpi_code: {school_year: {metric: cell}}}``
* ``SUBGROUP_BUILDERS[topic](frames, codes)`` — the same metrics broken out
  by student group, for the given district codes only (config districts +
  statewide; subgroup files for all 507 districts would grow the repo ~10x
  for districts nobody can route to):
  ``{dpi_code: {school_year: {dimension_id: {group_label: {metric: cell}}}}}``

``frames`` is ``{topic: {school_year: {member: DataFrame}}}`` (member = CSV
filename prefix before "_certified"; the attendance_dropouts ZIP yields
members "attendance" and "dropouts").

Row selection, verified against real files (see sources.py recon notes):
district rows are SCHOOL_NAME == "[Districtwide]", statewide rows are
DISTRICT_CODE == "0000" with SCHOOL_NAME == "[Statewide]"; both filtered to
GRADE_GROUP == "[All]". All-students rows use GROUP_BY == "All Students";
subgroup rows use the GROUP_BY labels in DIMENSIONS (drift across years:
"ELL Status" became "EL Status", and the value "ELL/LEP" became "EL" —
harmonized via GROUP_VALUE_ALIASES with a methodology note frontend-side).
"""

import openpyxl
import pandas as pd

from validate import SUPPRESSION_MARKERS

Cell = dict  # {"value": float | None, "suppressed": bool}

SUPPRESSED: Cell = {"value": None, "suppressed": True}

# dimension id -> GROUP_BY labels used across years/topics.
# "Disability Status" is the SwD/SwoD binary; the separate "Disability"
# GROUP_BY (by condition: Autism, ...) is deliberately not ingested.
DIMENSIONS: dict[str, list[str]] = {
    "race_ethnicity": ["Race/Ethnicity"],
    "econ_status": ["Economic Status"],
    "disability": ["Disability Status"],
    "el_status": ["EL Status", "ELL Status"],
}

# DPI renamed the EL category value around 2019; one population, one label.
GROUP_VALUE_ALIASES: dict[str, str] = {"ELL/LEP": "EL"}

# A redaction bucket, not a student group: rows whose GROUP_BY_VALUE hides
# which group they belong to. Skipped entirely.
SKIP_GROUP_VALUES: set[str] = {"[Data Suppressed]"}


def to_cell(raw) -> Cell:
    """Convert one raw CSV value to a cell. Suppression markers become
    null+suppressed; numbers become numbers; anything else throws."""
    if isinstance(raw, str):
        stripped = raw.strip()
        if stripped in SUPPRESSION_MARKERS:
            return {"value": None, "suppressed": True}
        raw = stripped
    try:
        return {"value": float(raw), "suppressed": False}
    except (TypeError, ValueError):
        raise ValueError(
            f"Unclassifiable cell value {raw!r}: not a number and not a known "
            f"suppression marker {sorted(SUPPRESSION_MARKERS)}. If DPI uses a "
            "new marker, add it to validate.SUPPRESSION_MARKERS deliberately."
        )


def _is_suppressed(raw) -> bool:
    return isinstance(raw, str) and raw.strip() in SUPPRESSION_MARKERS


def _num(raw, context: str) -> float:
    cell = to_cell(raw)
    if cell["suppressed"]:
        raise ValueError(f"{context}: unexpected suppressed value in a computed input")
    return cell["value"]


def _get_member(frames: dict, topic: str, member: str) -> dict[str, pd.DataFrame]:
    """{school_year: frame} for one member CSV of one topic. Throws if any
    year's ZIP lacked that member."""
    out = {}
    for year, members in frames[topic].items():
        if member not in members:
            raise ValueError(
                f"{topic} {year}: expected member CSV '{member}_certified_*' in ZIP, "
                f"got {sorted(members)}"
            )
        out[year] = members[member]
    return out


def _select(df: pd.DataFrame, year: str, topic: str, group_by: str,
            codes: set[str] | None = None) -> pd.DataFrame:
    """District-level rows (plus the one statewide row-group) for one
    GROUP_BY label. codes limits to specific districts (subgroup builds)."""
    for col in ("DISTRICT_CODE", "DISTRICT_NAME", "SCHOOL_NAME", "GROUP_BY"):
        if col not in df.columns:
            raise ValueError(f"{topic} {year}: expected column {col} missing "
                             f"from {list(df.columns)}")
    mask = (df["GROUP_BY"] == group_by) & (
        (df["SCHOOL_NAME"] == "[Districtwide]")
        | ((df["DISTRICT_CODE"] == "0000") & (df["SCHOOL_NAME"] == "[Statewide]"))
    )
    if "GRADE_GROUP" in df.columns:
        mask &= df["GRADE_GROUP"] == "[All]"
    if codes is not None:
        mask &= df["DISTRICT_CODE"].isin(codes)
    return df[mask]


# ---------------------------------------------------------------------------
# Per-topic metric extraction. Each function takes the rows for ONE
# (district, group) and returns {metric: cell} — or None to omit the entry.
# `strict` is True for all-students builds (a malformed file should fail the
# run) and False for subgroups, where DPI legitimately omits whole subjects
# or rows for tiny groups.
# ---------------------------------------------------------------------------

def _one_row(grp: pd.DataFrame, ctx: str) -> pd.Series:
    if len(grp) != 1:
        raise ValueError(f"{ctx}: expected exactly one row, got {len(grp)}")
    return grp.iloc[0]


def _enrollment_metrics(grp, ctx, strict):
    row = _one_row(grp, ctx)
    if pd.isna(row["STUDENT_COUNT"]):
        raise ValueError(f"{ctx}: empty STUDENT_COUNT")
    return {"total_enrollment": to_cell(row["STUDENT_COUNT"])}


def _dropout_metrics(grp, ctx, strict):
    row = _one_row(grp, ctx)
    for col in ("DROPOUT_RATE", "DROPOUT_COUNT"):
        if pd.isna(row[col]):
            raise ValueError(f"{ctx}: empty {col}")
    return {"dropout_rate": to_cell(row["DROPOUT_RATE"]),
            "dropout_count": to_cell(row["DROPOUT_COUNT"])}


def _chronic_absenteeism_metrics(grp, ctx, strict):
    row = _one_row(grp, ctx)
    if pd.isna(row["ABSENCE_RATE"]):
        raise ValueError(f"{ctx}: empty ABSENCE_RATE")
    return {"chronic_absenteeism_rate": to_cell(row["ABSENCE_RATE"])}


def _attendance_metrics(grp, ctx, strict):
    row = _one_row(grp, ctx)
    raw = row["ATTENDANCE_RATE"]
    # "--" = 0 possible days of attendance, division undefined (seen in
    # attendance_certified_2024-25.csv, PK/virtual rows). N/A, not privacy
    # suppression: omit the metric, don't fake a suppressed cell.
    if isinstance(raw, str) and raw.strip() == "--":
        return None
    return {"attendance_rate": to_cell(raw)}


# COMPLETION_STATUS wording drifts across years ("Completed - Regular" in
# 2009-10, "Completed - Regular High School Diploma" in 2024-25); the
# stable prefix identifies a regular-diploma completion in every year.
_REGULAR_PREFIX = "Completed - Regular"


def _graduation_metrics_for(suffix: str, include_counts: bool):
    """Regular-diploma completion metrics for one TIMEFRAME. The 4-year
    view keeps grad/cohort counts; the 5-/6-year views carry the rate only
    (the extra counts add little and would triple the metric pills). Note
    the year semantics match DPI's presentation: a file year's 5-year rate
    belongs to the cohort that was five years out as of that year."""
    def fn(grp, ctx, strict):
        if grp["COHORT"].nunique() > 1:
            raise ValueError(f"{ctx}: multiple {suffix} cohorts "
                             f"{sorted(grp['COHORT'].unique())}")
        cohort_raws = grp["COHORT_COUNT"].unique()
        if len(cohort_raws) != 1:
            raise ValueError(f"{ctx}: conflicting COHORT_COUNT values {cohort_raws}")
        cohort_raw = cohort_raws[0]

        regular = grp[grp["COMPLETION_STATUS"].str.startswith(_REGULAR_PREFIX)]
        if len(regular) > 1:
            raise ValueError(f"{ctx}: multiple regular-diploma rows")
        statuses_suppressed = bool((grp["COMPLETION_STATUS"] == "*").any())

        metrics: dict[str, Cell] = {}
        if include_counts:
            metrics[f"cohort_count_{suffix}"] = to_cell(cohort_raw)
        if len(regular) == 1:
            grad_raw = regular.iloc[0]["STUDENT_COUNT"]
            if include_counts:
                metrics[f"grad_count_{suffix}"] = to_cell(grad_raw)
            if _is_suppressed(grad_raw) or _is_suppressed(cohort_raw):
                metrics[f"grad_rate_{suffix}"] = dict(SUPPRESSED)
            else:
                rate = 100.0 * _num(grad_raw, ctx) / _num(cohort_raw, ctx)
                metrics[f"grad_rate_{suffix}"] = {"value": round(rate, 1),
                                                  "suppressed": False}
        elif statuses_suppressed:
            # Status breakdown redacted for privacy: rate unknowable.
            if include_counts:
                metrics[f"grad_count_{suffix}"] = dict(SUPPRESSED)
            metrics[f"grad_rate_{suffix}"] = dict(SUPPRESSED)
        else:
            # Statuses enumerated, none of them regular-diploma: a real zero
            # (tiny cohorts), not missing data.
            if include_counts:
                metrics[f"grad_count_{suffix}"] = {"value": 0.0, "suppressed": False}
            if _is_suppressed(cohort_raw):
                metrics[f"grad_rate_{suffix}"] = dict(SUPPRESSED)
            else:
                metrics[f"grad_rate_{suffix}"] = {"value": 0.0, "suppressed": False}
        return metrics
    return fn


def _ap_metrics(grp, ctx, strict):
    """AP participation and results from the AP_EXAM == '[All]' rollup
    (per-exam detail rows are deliberately not ingested)."""
    row = _one_row(grp, ctx)
    for col in ("STUDENTS_TESTED", "EXAM_COUNT", "PERCENT_3_OR_ABOVE"):
        if pd.isna(row[col]):
            raise ValueError(f"{ctx}: empty {col}")
    return {
        "students_tested": to_cell(row["STUDENTS_TESTED"]),
        "exam_count": to_cell(row["EXAM_COUNT"]),
        "pct_3_or_above": to_cell(row["PERCENT_3_OR_ABOVE"]),
    }


# TEST_SUBJECT -> metric name. 2024-25 adds Reading/Writing (ELA subscore)
# subjects we deliberately ignore; ELA itself exists in every year.
_FORWARD_SUBJECTS = {
    "ELA": "ela_prof_pct",
    "Mathematics": "math_prof_pct",
    "Science": "science_prof_pct",
    "Social Studies": "socstudies_prof_pct",
}

# Top-two performance categories under both cut-score regimes: 2015-16
# through 2022-23 use Proficient/Advanced, 2023-24 onward use
# Meeting/Advanced. One union set works because no year mixes regimes.
# The regimes themselves are NOT comparable — that's the
# cutscores-2023-24-forward comparability break in config/breaks.json.
_FORWARD_PROFICIENT = {"Proficient", "Advanced", "Meeting"}


def _forward_metrics(grp, ctx, strict):
    """Percent of students scoring in the top two performance categories,
    grades 3-8 combined per subject. The files carry per-grade rows only
    (no all-grades rollup), so the combined rate is summed from per-grade
    counts against DPI's own denominators (GROUP_COUNT, which includes
    No Test / No Score rows — matching DPI's published PERCENT_OF_GROUP
    semantics). If any contributing count is redacted the combined rate is
    unknowable and the metric is suppressed — never approximated."""
    metrics: dict[str, Cell] = {}
    for subject, metric in _FORWARD_SUBJECTS.items():
        srows = grp[grp["TEST_SUBJECT"] == subject]
        if srows.empty:
            continue  # subject not offered that year / for that group
        # A literal '*' TEST_RESULT row means the category breakdown itself
        # is redacted for that grade.
        breakdown_redacted = bool(srows["TEST_RESULT"].eq("*").any())
        prof_rows = srows[srows["TEST_RESULT"].isin(_FORWARD_PROFICIENT)]
        numerator_redacted = bool(prof_rows["STUDENT_COUNT"].map(_is_suppressed).any())

        denominator = 0.0
        denominator_redacted = False
        for grade, grows in srows.groupby("GRADE_LEVEL"):
            graws = {v for v in grows["GROUP_COUNT"].dropna()}
            if len(graws) != 1:
                raise ValueError(f"{ctx} {subject} grade {grade}: conflicting "
                                 f"GROUP_COUNT {sorted(graws)}")
            graw = graws.pop()
            if _is_suppressed(graw):
                denominator_redacted = True
            else:
                denominator += _num(graw, ctx)

        if breakdown_redacted or numerator_redacted or denominator_redacted:
            metrics[metric] = dict(SUPPRESSED)
        elif denominator > 0:
            numerator = sum(_num(v, ctx) for v in prof_rows["STUDENT_COUNT"])
            metrics[metric] = {"value": round(100.0 * numerator / denominator, 1),
                               "suppressed": False}
        # denominator 0 with nothing redacted: no students, omit.
    return metrics or None


def _preact_metrics(grp, ctx, strict):
    """PreACT Secure composite averages, one metric per grade level (9/10).
    Same AVERAGE_SCORE semantics as the ACT files."""
    metrics: dict[str, Cell] = {}
    for grade in ("9", "10"):
        rows = grp[(grp["TEST_SUBJECT"] == "Composite") & (grp["GRADE_LEVEL"] == grade)]
        if rows.empty:
            continue  # a district without that grade simply lacks the metric
        numeric = {v for v in rows["AVERAGE_SCORE"].dropna() if not _is_suppressed(v)}
        if len(numeric) > 1:
            raise ValueError(f"{ctx} grade {grade}: conflicting AVERAGE_SCORE "
                             f"{sorted(numeric)}")
        metric = f"composite_avg_gr{grade}"
        if numeric:
            metrics[metric] = to_cell(numeric.pop())
        elif rows["AVERAGE_SCORE"].map(_is_suppressed).any():
            metrics[metric] = dict(SUPPRESSED)
    return metrics or None


# TEST_SUBJECT -> metric name. "Composite" exists in every year 2014-15
# onward (2014-15/2015-16 also carry a "Combined" subject we ignore).
_ACT_SUBJECTS = {
    "Composite": "composite_avg",
    "English": "english_avg",
    "Mathematics": "math_avg",
    "Reading": "reading_avg",
    "Science": "science_avg",
}


def _act_metrics(grp, ctx, strict):
    metrics: dict[str, Cell] = {}
    for subject, metric in _ACT_SUBJECTS.items():
        rows = grp[grp["TEST_SUBJECT"] == subject]
        if rows.empty:
            if strict:
                raise ValueError(f"{ctx}: no rows for subject {subject}")
            continue  # tiny subgroups can lack whole subjects
        numeric = {v for v in rows["AVERAGE_SCORE"].dropna() if not _is_suppressed(v)}
        if len(numeric) > 1:
            raise ValueError(f"{ctx} {subject}: conflicting AVERAGE_SCORE "
                             f"{sorted(numeric)}")
        if numeric:
            metrics[metric] = to_cell(numeric.pop())
        elif rows["AVERAGE_SCORE"].map(_is_suppressed).any():
            metrics[metric] = dict(SUPPRESSED)
        # else: only 'No Test' rows, no average exists -> omit.

    comp = grp[grp["TEST_SUBJECT"] == "Composite"]
    if comp.empty:
        return metrics or None
    group_raws = {v for v in comp["GROUP_COUNT"].dropna()}
    if len(group_raws) != 1:
        raise ValueError(f"{ctx}: expected one GROUP_COUNT, got {sorted(group_raws)}")
    group_raw = group_raws.pop()
    no_test = comp[comp["TEST_RESULT"] == "No Test"]
    if len(no_test) > 1:
        raise ValueError(f"{ctx}: multiple Composite 'No Test' rows")
    # A missing 'No Test' row means zero non-participants ONLY when the
    # result rows are actually enumerated. If every Composite result row is
    # redacted ('*'), the number of non-testers is unknowable — computing
    # 100% participation there would fabricate a statistic.
    results_all_suppressed = bool(comp["TEST_RESULT"].eq("*").all())
    no_test_raw = no_test.iloc[0]["STUDENT_COUNT"] if len(no_test) else "0"

    if _is_suppressed(group_raw) or _is_suppressed(no_test_raw) or (
            len(no_test) == 0 and results_all_suppressed):
        metrics["participation_pct"] = dict(SUPPRESSED)
        metrics["tested_count"] = dict(SUPPRESSED)
    else:
        enrolled = _num(group_raw, ctx)
        tested = enrolled - _num(no_test_raw, ctx)
        metrics["tested_count"] = {"value": tested, "suppressed": False}
        metrics["participation_pct"] = {
            "value": round(100.0 * tested / enrolled, 1),
            "suppressed": False,
        }
    return metrics


# ---------------------------------------------------------------------------
# Topic definitions: where the rows come from and how to slice them.
# prep narrows the member frame before row-group selection (ACT's TEST_GROUP
# filter, graduation's 4-year timeframe). multi_row topics have one row-GROUP
# per district+group; the rest must be exactly one row.
# ---------------------------------------------------------------------------

_TOPIC_SOURCES = {
    "enrollment": [("enrollment", "enrollment", None, _enrollment_metrics)],
    "dropouts": [("dropouts", "dropouts", None, _dropout_metrics)],
    "absenteeism": [
        ("absenteeism", "absenteeism", None, _chronic_absenteeism_metrics),
        ("dropouts", "attendance", None, _attendance_metrics),
    ],
    "graduation": [
        ("graduation", "hs_completion",
         lambda df: df[df["TIMEFRAME"] == "4-Year rate"],
         _graduation_metrics_for("4yr", include_counts=True)),
        # 5-/6-year rows exist 2013-14 onward; earlier years just lack them
        # (the trailing True marks the source optional for empty years).
        ("graduation", "hs_completion",
         lambda df: df[df["TIMEFRAME"] == "5-Year rate"],
         _graduation_metrics_for("5yr", include_counts=False), True),
        ("graduation", "hs_completion",
         lambda df: df[df["TIMEFRAME"] == "6-Year rate"],
         _graduation_metrics_for("6yr", include_counts=False), True),
    ],
    "act": [(
        "act", "act_statewide",
        lambda df: df[df["TEST_GROUP"] == "ACT"],
        _act_metrics,
    )],
    "preact": [(
        "preact", "preact_secure_statewide",
        lambda df: df[df["TEST_GROUP"] == "PreACT"],
        _preact_metrics,
    )],
    # Grades 3-8 only (Science/Social Studies also carry grade-10 rows in
    # some years) for a stable definition across the whole series.
    "forward": [(
        "forward", "forward",
        lambda df: df[(df["TEST_GROUP"] == "Forward")
                      & (df["GRADE_LEVEL"].isin(["3", "4", "5", "6", "7", "8"]))],
        _forward_metrics,
    )],
    "ap": [(
        "ap", "ap",
        lambda df: df[df["AP_EXAM"] == "[All]"],
        _ap_metrics,
    )],
}


def _select_schools(df: pd.DataFrame, year: str, topic: str, codes: set[str]) -> pd.DataFrame:
    """Per-school rows for the given districts, all-students only. School
    rows carry their school type in GRADE_GROUP (Elementary School, High
    School, ...) rather than '[All]', so no GRADE_GROUP filter applies;
    bracket pseudo-schools ([Districtwide], [Statewide], any future ones)
    are excluded wholesale."""
    for col in ("DISTRICT_CODE", "SCHOOL_CODE", "SCHOOL_NAME", "GROUP_BY"):
        if col not in df.columns:
            raise ValueError(f"{topic} {year}: expected column {col} missing "
                             f"from {list(df.columns)}")
    mask = ((df["GROUP_BY"] == "All Students")
            & df["DISTRICT_CODE"].isin(codes)
            & ~df["SCHOOL_NAME"].str.startswith("["))
    return df[mask]


def _build_schools(topic: str, frames: dict, codes: set[str]) -> dict:
    """{district_code: {school_code: {"name", "type", "years": {year:
    {metric: cell}}}}} — the district-file metrics per school. strict=False
    like subgroups: schools legitimately lack subjects, timeframes, and
    whole topics (no ACT rows in an elementary school)."""
    out: dict[str, dict] = {}
    for src_topic, member, prep, metrics_fn, *rest in _TOPIC_SOURCES[topic]:
        member_frames = _get_member(frames, src_topic, member)
        for year in sorted(member_frames):
            df = member_frames[year]
            sel = _select_schools(prep(df) if prep else df, year, topic, codes)
            for (dcode, scode), grp in sel.groupby(["DISTRICT_CODE", "SCHOOL_CODE"]):
                ctx = f"{topic} {year} district {dcode} school {scode}"
                metrics = metrics_fn(grp, ctx, strict=False)
                if not metrics:
                    continue
                school = out.setdefault(dcode, {}).setdefault(
                    scode, {"name": None, "type": None, "years": {}})
                # Years iterate sorted, so the newest name/type win —
                # school names drift across years like district names do.
                school["name"] = grp.iloc[0]["SCHOOL_NAME"]
                school["type"] = grp.iloc[0]["GRADE_GROUP"]
                school["years"].setdefault(year, {}).update(metrics)
    return out


# ---------------------------------------------------------------------------
# Non-WISEdash xlsx builders. These parse DPI program-office spreadsheets
# (see sources.XLSX_FILES) instead of the WISEdash frame machinery, but
# emit the exact same {code: {year: {metric: cell}}} shape.
# ---------------------------------------------------------------------------

# Expected column headers after YEAR/DIST NO/DISTRICT NAME, in file order.
_OE_COLUMNS = ["PUPIL TRANSFERS IN", "PUPIL TRANSFERS OUT", "NET PUPIL TRANSFERS",
               "AID TRANSFERS IN", "AID TRANSFERS OUT", "NET AID TRANSFERS"]
_OE_METRICS = ["pupils_in", "pupils_out", "net_pupils",
               "aid_in", "aid_out", "net_aid"]


def build_open_enrollment(xlsx_paths: dict[str, "object"]) -> dict:
    """Open enrollment pupil/aid transfers per district-year from the
    program office's annual xlsx. Cells are used verbatim: DPI sometimes
    redacts the NET column while leaving a component visible — computing
    the net ourselves would un-redact what DPI redacted, so we never do.
    Rows are matched by a 4-digit DISTRICT NO; the Totals row and footnote
    rows lack one and fall away. No statewide row exists in these files."""
    out: dict[str, dict] = {}
    for year, path in sorted(xlsx_paths.items()):
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        header_idx = next(
            (i for i, r in enumerate(rows)
             # 2018-19 titles the column 'Year' instead of 'YEAR'.
             if str(r[0]).strip().upper() == "YEAR"
             and isinstance(r[1], str) and r[1].startswith("DIST")),
            None)
        if header_idx is None:
            raise ValueError(f"open_enrollment {year}: no header row (YEAR / DIST* NO) "
                             f"found in {path}")
        header = [str(c).strip() if c else "" for c in rows[header_idx]]
        if header[3:9] != _OE_COLUMNS:
            raise ValueError(f"open_enrollment {year}: unexpected columns {header[3:9]} "
                             f"(expected {_OE_COLUMNS}) — DPI changed the layout.")
        seen = 0
        for r in rows[header_idx + 1:]:
            code = r[1]
            if not (isinstance(code, str) and len(code) == 4 and code.isdigit()):
                continue  # Totals row, footnote, blank tail
            # A district with no open enrollment activity gets an all-blank
            # row (Washington Island 2016-17: five blanks + a 0 net aid).
            # Blank is not zero and not suppression — omit the year.
            if all(v is None for v in r[3:8]):
                seen += 1
                continue
            metrics = {}
            for metric, raw in zip(_OE_METRICS, r[3:9]):
                metrics[metric] = to_cell(raw)
            out.setdefault(code, {})[year] = metrics
            seen += 1
        if seen < 400:
            raise ValueError(f"open_enrollment {year}: only {seen} district rows "
                             "parsed — expected ~421; the file layout likely moved.")
    return out


def _fiscal_label(y: int) -> str:
    return f"{y}-{str(y + 1)[-2:]}"


def _fin_num(v, ctx: str):
    """Finance workbook cell -> float, None for blank-ish (None / ''),
    error otherwise — these workbooks have stray whitespace-only cells
    but are never privacy-redacted."""
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        # '' = blank; '#DIV/0!' etc. = the workbook's own formula failed
        # (zero-membership years) — no value exists, same as blank.
        if not v or v.startswith("#"):
            return None
    try:
        return float(v)
    except (TypeError, ValueError):
        raise ValueError(f"{ctx}: unparseable finance cell {v!r}")


def build_finance(paths: dict[str, "object"], compcost_end: str) -> dict:
    """cost_per_member + revenue_limit_per_member per district-year from
    the two SFS longitudinal workbooks. Both key districts by UNPADDED
    numeric code — zero-filled here to match the repo's codes. Finance
    data is never redacted, so any unparseable cell is an error, not
    suppression. compcost_end (e.g. '2024-25', parsed from the served
    filename by refresh.py) caps the cost series: the workbook carries a
    year group beyond DPI's published audited range — current-year
    unaudited figures we refuse to ship.

    Returns (data, names): the finance history reaches back before the
    WISEdash files start (1993-94 vs 2005-06), so districts dissolved
    before 2005 appear ONLY here — their names come from these workbooks
    (used as lowest-priority entries in the refresh name map)."""
    out: dict[str, dict] = {}
    names: dict[str, str] = {}

    # --- Comparative cost per member (2008-09+). DATA sheet: header row
    # with DISTRICT_NAME then repeating groups starting at each
    # FISCAL_YEAR column: member + cost-category dollar columns (category
    # count drifts at DPI's 2023 recalculation; we sum whatever
    # categories each group carries and divide by member).
    wb = openpyxl.load_workbook(paths["compcost"], data_only=True, read_only=True)
    try:
        ws = wb["DATA"]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    groups = []  # (year_col, member_col, [cost_cols])
    for i, name in enumerate(header):
        # Header case drifts mid-file: 'FISCAL_YEAR' for the early year
        # groups, 'fiscal_year' for later ones.
        if name.upper() == "FISCAL_YEAR":
            groups.append({"year": i, "member": None, "costs": []})
        elif groups:
            if name.lower() == "member":
                groups[-1]["member"] = i
            elif name:
                groups[-1]["costs"].append(i)
    if not groups or any(g["member"] is None or not g["costs"] for g in groups):
        raise ValueError("finance compcost: FISCAL_YEAR/member/cost header "
                         f"groups not found in {header[:12]}...")
    for r in rows[1:]:
        code_raw = r[0]
        if not (isinstance(code_raw, (int, float)) and int(code_raw) > 0):
            continue  # header filler rows / footers
        code = str(int(code_raw)).zfill(4)
        if isinstance(r[1], str) and r[1].strip():
            names[code] = r[1].strip()
        for g in groups:
            ctx = f"finance compcost district {code}"
            year_raw = _fin_num(r[g["year"]], ctx)
            member = _fin_num(r[g["member"]], ctx)
            if year_raw is None or not member:
                continue  # district didn't exist / no membership that year
            costs = [_fin_num(r[i], ctx) for i in g["costs"]]
            if any(c is None for c in costs):
                continue  # year group not yet populated for this district
            year = _fiscal_label(int(year_raw))
            if year > compcost_end:
                continue  # beyond DPI's published audited range
            cell = {"value": round(sum(costs) / member), "suppressed": False}
            out.setdefault(code, {}).setdefault(year, {})["cost_per_member"] = cell

    # --- Revenue limit per member (1993-94+). Data sheet: row0 = year
    # labels like '93-94' (each spanning a 3-column group), row1 =
    # sub-headers ending in the per-member column, row2 = junk, data
    # after. The per-member value is the 3rd column of each triple.
    wb = openpyxl.load_workbook(paths["revenue_limit"], data_only=True, read_only=True)
    try:
        ws = next(wb[s] for s in wb.sheetnames if s.strip().lower() == "data")
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    year_row = rows[0]
    triples = []  # (year_label, per_member_col)
    col = 2
    while col + 2 < len(year_row):
        label = str(year_row[col]).strip() if year_row[col] is not None else ""
        if not label or "-" not in label:
            break
        start, _end = label.split("-")
        start_i = int(start)
        full = 1900 + start_i if start_i >= 50 else 2000 + start_i
        triples.append((_fiscal_label(full), col + 2))
        col += 3
    if len(triples) < 25:
        raise ValueError(f"finance revenue_limit: only {len(triples)} year triples "
                         "parsed from the Data sheet header — layout moved.")
    for r in rows[3:]:
        code_raw = r[0]
        if not (isinstance(code_raw, (int, float)) and int(code_raw) > 0):
            continue  # State Totals / filler rows
        code = str(int(code_raw)).zfill(4)
        if code not in names and isinstance(r[1], str) and r[1].strip():
            names[code] = r[1].strip()
        for year, pcol in triples:
            v = _fin_num(r[pcol] if pcol < len(r) else None,
                         f"finance revenue_limit district {code} {year}")
            if not v:
                continue  # merged/closed district years carry blanks or 0
            cell = {"value": round(v), "suppressed": False}
            out.setdefault(code, {}).setdefault(year, {})["revenue_limit_per_member"] = cell

    if len(out) < 400:
        raise ValueError(f"finance: only {len(out)} districts parsed — expected ~430.")
    return out, names


# Referendum event fields kept from the WiSFPR payload. Not cell-shaped
# data (an event list, not year->metric), so it ships as its own file
# family (data/referenda/) with its own schema.
def build_referenda(payload: dict, codes: set[str]) -> dict:
    """{code: [event, ...]} for the given districts, oldest first. Every
    event must carry the fields the frontend renders — fail fast on any
    reshape of the endpoint's JSON."""
    rows = payload.get("Data")
    if not isinstance(rows, list) or len(rows) < 3000:
        raise ValueError("referenda: endpoint payload missing Data or "
                         f"implausibly small ({type(rows)} / {len(rows) if isinstance(rows, list) else 'n/a'}) "
                         "— did the date-range default kick in?")
    out: dict[str, list] = {}
    for r in rows:
        code = r.get("AgencyCode")
        if code not in codes:
            continue
        for field in ("VoteDate", "ReferendumType", "ReferendumTypeCode",
                      "Amount", "ReferendumStatus"):
            if field not in r:
                raise ValueError(f"referenda: row missing {field}: {sorted(r)}")
        out.setdefault(code, []).append({
            "vote_date": r["VoteDate"][:10],
            "type_code": r["ReferendumTypeCode"],
            "type": r["ReferendumType"],
            "amount": r["Amount"],
            "brief": (r.get("BriefDescription") or "").strip(),
            "yes_votes": r.get("YesVotes"),
            "no_votes": r.get("NoVotes"),
            "status": r["ReferendumStatus"],
        })
    for events in out.values():
        events.sort(key=lambda e: e["vote_date"])
    return out


def source_topics(topic: str) -> set[str]:
    """Which source topics' frames a topic's builders read (absenteeism
    reads the dropouts ZIP's attendance member). Lets refresh.py load
    frames lazily and release them once no later topic needs them."""
    return {src for src, *_ in _TOPIC_SOURCES[topic]}


# Row prunes applied at CSV-load time purely to keep peak memory sane —
# the forward files run ~1M rows per year and the full corpus no longer
# fits in memory unpruned. A prune MUST keep a superset of every row any
# builder (all-students, subgroup, school-level) selects; dropping a row
# a builder would use corrupts output silently. Only provably-unused rows
# go: alternate-assessment DLM, grades outside 3-8, ELA subscore subjects,
# and GROUP_BY dimensions we deliberately don't ingest (Gender, Migrant).
LOAD_PRUNES = {
    "forward": lambda df: df[
        (df["TEST_GROUP"] == "Forward")
        & df["GRADE_LEVEL"].isin(["3", "4", "5", "6", "7", "8"])
        & df["TEST_SUBJECT"].isin(list(_FORWARD_SUBJECTS))
        & df["GROUP_BY"].isin(
            ["All Students"] + [l for ls in DIMENSIONS.values() for l in ls])
    ],
}


def _build_all_students(topic: str, frames: dict) -> dict:
    out: dict[str, dict] = {}
    for src_topic, member, prep, metrics_fn, *rest in _TOPIC_SOURCES[topic]:
        optional = rest[0] if rest else False
        for year, df in _get_member(frames, src_topic, member).items():
            sel = _select(prep(df) if prep else df, year, topic, "All Students")
            if sel.empty:
                # Optional sources legitimately have no rows in some years
                # (e.g. 5-/6-year graduation timeframes before 2013-14).
                if optional:
                    continue
                raise ValueError(f"{topic} {year}: row selection returned nothing")
            for code, grp in sel.groupby("DISTRICT_CODE"):
                ctx = f"{topic} {year} district {code}"
                metrics = metrics_fn(grp, ctx, strict=True)
                if metrics:
                    out.setdefault(code, {}).setdefault(year, {}).update(metrics)
    return out


def _build_subgroups(topic: str, frames: dict, codes: set[str]) -> dict:
    out: dict[str, dict] = {}
    for src_topic, member, prep, metrics_fn, *rest in _TOPIC_SOURCES[topic]:
        for year, df in _get_member(frames, src_topic, member).items():
            narrowed = prep(df) if prep else df
            for dim, labels in DIMENSIONS.items():
                present = [l for l in labels if (narrowed["GROUP_BY"] == l).any()]
                for label in present:
                    sel = _select(narrowed, year, topic, label, codes=codes)
                    for (code, raw_value), grp in sel.groupby(
                            ["DISTRICT_CODE", "GROUP_BY_VALUE"]):
                        if raw_value in SKIP_GROUP_VALUES:
                            continue
                        value = GROUP_VALUE_ALIASES.get(raw_value, raw_value)
                        ctx = f"{topic} {year} district {code} {dim}/{value}"
                        metrics = metrics_fn(grp, ctx, strict=False)
                        if metrics:
                            (out.setdefault(code, {})
                                .setdefault(year, {})
                                .setdefault(dim, {})
                                .setdefault(value, {})
                                .update(metrics))
    return out


def build_enrollment(frames):
    """total_enrollment from enrollment_certified STUDENT_COUNT."""
    return _build_all_students("enrollment", frames)


def build_dropouts(frames):
    """dropout_rate (%, grades 7-12) and dropout_count from the dropouts
    member of the attendance_dropouts ZIP."""
    return _build_all_students("dropouts", frames)


def build_absenteeism(frames):
    """chronic_absenteeism_rate (ESSA/STATE ABSENCE_RATE) plus
    attendance_rate from the attendance member of the attendance_dropouts
    ZIP downloaded for the dropouts topic."""
    return _build_all_students("absenteeism", frames)


def build_graduation(frames):
    """4-year cohort regular-diploma graduation: grad_rate_4yr (%),
    grad_count_4yr, cohort_count_4yr."""
    return _build_all_students("graduation", frames)


def build_act(frames):
    """Grade-11 census ACT, TEST_GROUP == 'ACT' only: per-subject averages,
    participation_pct and tested_count from the Composite row-group."""
    return _build_all_students("act", frames)


def build_preact(frames):
    """Grades 9-10 PreACT Secure composite averages (TEST_GROUP == 'PreACT')."""
    return _build_all_students("preact", frames)


def build_forward(frames):
    """Grades 3-8 Forward Exam: percent in the top two performance
    categories per subject, grades combined (TEST_GROUP == 'Forward')."""
    return _build_all_students("forward", frames)


def build_ap(frames):
    """AP participation and results from the AP_EXAM == '[All]' rollup."""
    return _build_all_students("ap", frames)


BUILDERS = {
    "act": build_act,
    "forward": build_forward,
    "preact": build_preact,
    "ap": build_ap,
    "graduation": build_graduation,
    "dropouts": build_dropouts,
    "absenteeism": build_absenteeism,
    "enrollment": build_enrollment,
}

SUBGROUP_BUILDERS = {
    topic: (lambda frames, codes, _t=topic: _build_subgroups(_t, frames, codes))
    for topic in _TOPIC_SOURCES
}

SCHOOL_BUILDERS = {
    topic: (lambda frames, codes, _t=topic: _build_schools(_t, frames, codes))
    for topic in _TOPIC_SOURCES
}
